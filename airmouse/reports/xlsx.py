"""Excel report export."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font


def export_xlsx(stats: dict, output_path: Path) -> None:
    if not stats:
        raise ValueError("No statistics available for export")

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    ws["A1"] = "AirMouse Gesture Report"
    ws["A1"].font = Font(size=14, bold=True)

    row = 3
    for key in ("duration", "frames", "fps", "avg_confidence", "avg_latency_ms", "device"):
        ws[f"A{row}"] = key
        ws[f"B{row}"] = str(stats.get(key, "-"))
        ws[f"A{row}"].font = Font(bold=True)
        row += 1

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 24

    series_sheet = wb.create_sheet("Series")
    series_sheet["A1"] = "idx"
    series_sheet["B1"] = "fps"
    series_sheet["C1"] = "confidence"
    series_sheet["D1"] = "latency_ms"

    fps_series = stats.get("fps_series", [])
    conf_series = stats.get("confidence_series", [])
    lat_series = stats.get("latency_series_ms", [])
    max_len = max(len(fps_series), len(conf_series), len(lat_series))

    for idx in range(max_len):
        row_idx = idx + 2
        series_sheet[f"A{row_idx}"] = idx + 1
        series_sheet[f"B{row_idx}"] = float(fps_series[idx]) if idx < len(fps_series) else None
        series_sheet[f"C{row_idx}"] = float(conf_series[idx]) if idx < len(conf_series) else None
        series_sheet[f"D{row_idx}"] = float(lat_series[idx]) if idx < len(lat_series) else None

    if max_len > 1:
        chart = LineChart()
        chart.title = "Runtime dynamics"
        chart.y_axis.title = "value"
        chart.x_axis.title = "step"
        data = Reference(series_sheet, min_col=2, min_row=1, max_col=4, max_row=max_len + 1)
        cats = Reference(series_sheet, min_col=1, min_row=2, max_row=max_len + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        series_sheet.add_chart(chart, "F2")

    wb.save(output_path)
